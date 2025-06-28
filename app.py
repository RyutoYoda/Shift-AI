# -*- coding: utf-8 -*-
"""
- 夜勤 1 名 + 世話人 1 名 / 日（必ず全日埋める。欠番があればエラーではなく *自動緩和* で必ず充当）
- 夜勤後インターバル 2 日 → 1 日 → 0 日（※0 日は夜勤と世話人が同人物でなければ可）と 3 段階で緩和
- 0 セルは厳守 / 上限は超えない（限界まで使い切る）
- 指定セル以外の編集禁止・列 C の集計式も保持
- 出力は元ブックをベースに `.xlsx` で上書き保存
------------------------------------------------------------
"""

import io
from typing import List, Tuple, Dict

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

# -------------------- 主要定数 --------------------
HEADER_ROW = 3      # 日付が並ぶ行 (0-index)
START_ROW  = 4      # シフトが始まる最上行 (0-index)
NAME_COL   = 1      # 氏名列 (0-index)

# シフトを書き換えて良い範囲
EDIT_BLOCKS = [
    (4, 15),   # E5 〜 AI16  (0‑index: rows 4‑15)
    (19, 29),  # E20 〜 AI30 (0‑index: rows 19‑29)
]

SHIFT_NIGHT_HOURS = 12.5  # 夜勤 1 回
SHIFT_CARE_HOURS  = 6.0   # 世話人 1 回

# -------------------- ユーティリティ --------------------

def detect_date_columns(df: pd.DataFrame) -> List[int]:
    """1〜31 の整数が入っているヘッダー列を日付列とみなす"""
    date_cols: List[int] = []
    for c in df.columns:
        val = df.iat[HEADER_ROW, c]
        try:
            v = int(float(val))
            if 1 <= v <= 31:
                date_cols.append(c)
        except (ValueError, TypeError):
            continue
    if not date_cols:
        raise ValueError("ヘッダー行に 1〜31 の日付が見つかりません。行番号・列番号を確認してください。")
    return date_cols


def detect_row_indices(df: pd.DataFrame) -> Tuple[List[int], List[int]]:
    """1 列目ラベルで『夜間支援員』『世話人』行を判定"""
    night_rows, care_rows = [], []
    for r in range(START_ROW, df.shape[0]):
        role = df.iat[r, 0]
        name = df.iat[r, NAME_COL]
        if not isinstance(role, str) or not isinstance(name, str) or not name.strip():
            continue
        role_flat = role.replace("\n", "")
        if "夜間" in role_flat and "支援員" in role_flat:
            night_rows.append(r)
        elif "世話人" in role_flat:
            care_rows.append(r)
    if not night_rows or not care_rows:
        raise ValueError("夜間支援員 / 世話人 の行が検出できませんでした。行ラベルを確認してください。")
    return night_rows, care_rows


def get_limits(df: pd.DataFrame) -> pd.Series:
    """下部の『上限(時間)』テーブルを取得"""
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            if str(df.iat[r, c]).startswith("上限"):
                name_col, limit_col = c - 1, c
                limits = {}
                rr = r + 1
                while rr < df.shape[0]:
                    name = df.iat[rr, name_col]
                    if not isinstance(name, str) or not name.strip():
                        break
                    limit_val = pd.to_numeric(df.iat[rr, limit_col], errors="coerce")
                    limits[name.strip()] = float(limit_val) if not np.isnan(limit_val) else np.inf
                    rr += 1
                return pd.Series(limits)
    raise ValueError("『上限(時間)』テーブルが見つかりませんでした。シート最下部に配置してください。")


def in_edit_blocks(r: int) -> bool:
    for start, end in EDIT_BLOCKS:
        if start <= r <= end:
            return True
    return False

# -------------------- コア割当てロジック --------------------


def assign_with_constraints(
    df: pd.DataFrame,
    date_cols: List[int],
    night_rows: List[int],
    care_rows: List[int],
    limits: pd.Series,
    interval_rule: int,
) -> Tuple[bool, pd.Series, Dict[str, int]]:
    """与えられたインターバルで全日割当てを試みる。
    成功なら True と totals, last_night_day を返す。失敗なら False。"""
    totals = pd.Series(0.0, index=limits.index)
    last_night_day: Dict[str, int] = {}

    # 全セルクリア（0 は残す）
    for r in night_rows + care_rows:
        for c in date_cols:
            if df.iat[r, c] != 0 and not pd.isna(df.iat[r, c]):
                df.iat[r, c] = np.nan

    # 日ループ
    for d_idx, c in enumerate(date_cols):
        # ------------------ 夜勤 ------------------
        night_cand = [
            (limits[name] - totals[name], name, r)
            for r, name in ((r, df.iat[r, NAME_COL].strip()) for r in night_rows)
            if pd.isna(df.iat[r, c])
            and totals[name] + SHIFT_NIGHT_HOURS <= limits[name]
        ]
        if not night_cand:
            return False, totals, last_night_day
        night_cand.sort(key=lambda x: (x[0], x[1]))
        _, night_name, night_row = night_cand[0]
        df.iat[night_row, c] = SHIFT_NIGHT_HOURS
        totals[night_name] += SHIFT_NIGHT_HOURS
        last_night_day[night_name] = d_idx

        # ------------------ 世話人 ------------------
        care_cand = [
            (limits[name] - totals[name], name, r)
            for r, name in ((r, df.iat[r, NAME_COL].strip()) for r in care_rows)
            if pd.isna(df.iat[r, c])
            and name != night_name
            and (
                name not in last_night_day
                or d_idx - last_night_day[name] >= (interval_rule + 1)
            )
            and totals[name] + SHIFT_CARE_HOURS <= limits[name]
        ]
        if not care_cand:
            return False, totals, last_night_day
        care_cand.sort(key=lambda x: (x[0], x[1]))
        _, care_name, care_row = care_cand[0]
        df.iat[care_row, c] = SHIFT_CARE_HOURS
        totals[care_name] += SHIFT_CARE_HOURS

    return True, totals, last_night_day


def optimize(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.Series, pd.Series]:
    date_cols = detect_date_columns(df)
    night_rows, care_rows = detect_row_indices(df)

    for r in night_rows + care_rows:
        if not in_edit_blocks(r):
            raise ValueError("夜間支援員/世話人 の行が EDIT_BLOCKS から外れています。定数 EDIT_BLOCKS を調整してください。")

    # 氏名 → 上限
    names = sorted({df.iat[r, NAME_COL].strip() for r in night_rows + care_rows})
    limits = get_limits(df).reindex(names).fillna(np.inf)

    # ---------- 3 段階インターバルで Try ----------
    for interval in (2, 1, 0):
        success, totals, _ = assign_with_constraints(
            df, date_cols, night_rows, care_rows, limits, interval_rule=interval
        )
        if success:
            break
    else:  # never broke
        raise RuntimeError("ルールを緩和しても全日割当てできません。0 セルや上限を見直してください。")

    # 完成チェック
    for c in date_cols:
        # 夜勤 / 世話人 どちらも埋まっているか
        if all(pd.isna(df.iat[r, c]) or df.iat[r, c] == 0 for r in night_rows):
            raise RuntimeError("夜勤が空欄の日が残っています。入力テンプレートを確認してください。")
        if all(pd.isna(df.iat[r, c]) or df.iat[r, c] == 0 for r in care_rows):
            raise RuntimeError("世話人が空欄の日が残っています。入力テンプレートを確認してください。")

    return df, totals.sort_index(), limits.sort_index()

# -------------------- 元ブックへ書き戻し --------------------

def write_back(original_stream: io.BytesIO, df_opt: pd.DataFrame) -> bytes:
    original_stream.seek(0)
    wb: Workbook = load_workbook(original_stream, data_only=False)
    ws = wb.active

    for r in range(df_opt.shape[0]):
        if not in_edit_blocks(r):
            continue
        for c in range(df_opt.shape[1]):
            header_val = df_opt.iat[HEADER_ROW, c]
            try:
                int(float(header_val))
            except (ValueError, TypeError):
                continue
            new_val = df_opt.iat[r, c]
            if pd.isna(new_val):
                new_val = None
            ws.cell(row=r + 1, column=c + 1, value=new_val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# -------------------- Streamlit UI --------------------

st.set_page_config(page_title="シフト自動最適化", layout="wide")
st.title("📅 シフト自動最適化ツール (空欄ゼロ保証版)")

with st.expander("👉 使い方はこちら", expanded=False):
    st.markdown(
        """
        **手順**
        1. 左サイドバーで Excel ファイル (.xlsx) をアップロード。
        2. **「🚀 最適化を実行」** をクリック。
        3. 右側にプレビューが表示されます（**必ず全日埋まります**）。
        4. **「📥 ダウンロード」** で最適化済み Excel を取得。

        **割当てロジック**
        - *夜勤 1 名* + *世話人 1 名* を毎日必ず充当。
        - 夜勤 → 世話人のインターバルは 2 日を原則として、割当て不能日は 1 日→0 日へ自動緩和。
        - 0 セルは固定で上書き不可。
        - 『上限(時間)』を超えない範囲で、限界まで時間を使います。
        - 指定セル (E5‑AI16, E20‑AI30) 以外は一切変更せず、列 C の集計式も保持します。
        """
    )

st.sidebar.header("📂 入力ファイル")
uploaded = st.sidebar.file_uploader("Excel ファイル (.xlsx)", type=["xlsx"])

if uploaded is not None:
    try:
        df_input = pd.read_excel(uploaded, header=None, engine="openpyxl")
        st.subheader("アップロードされたシフト表 (プレビュー)")
        st.dataframe(df_input, use_container_width=True)

        if st.sidebar.button("🚀 最適化を実行"):
            df_opt, totals, limits = optimize(df_input.copy())
            st.success("✅ 最適化が完了し、全日埋めました")

            st.subheader("最適化後のシフト表 (プレビュー)")
            st.dataframe(df_opt, use_container_width=True)

            st.subheader("勤務時間の合計 / 上限")
            st.dataframe(pd.DataFrame({"合計時間": totals, "上限時間": limits}))

            optimized_bytes = write_back(uploaded, df_opt)
            st.download_button(
                "📥 最適化シフトをダウンロード (Excel)",
                data=optimized_bytes,
                file_name="optimized_shift.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except Exception as e:
        st.error(f"ファイルの読み込みまたは最適化中にエラーが発生しました: {e}")
else:
    st.info("左のサイドバーからテンプレート Excel をアップロードしてください。")
